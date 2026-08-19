import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# ==============================================================================
# PORTFÖY AYARLARINIZ
# ==============================================================================
ay_sayisi = 1  
aylik_odeme = 5000.0
toplam_anapara = float(aylik_odeme * ay_sayisi)

fon_tanimlari = {
    'GEL': {'ad': 'Para Piyasası Emeklilik Yatırım Fonu', 'agirlik': 0.20, 'giris_fiyati': 0.436406},
    'GEH': {'ad': 'Hisse Senedi Emeklilik Yatırım Fonu', 'agirlik': 0.30, 'giris_fiyati': 2.779911},
    'EMY': {'ad': 'Altın Emeklilik Yatırım Fonu', 'agirlik': 0.20, 'giris_fiyati': 0.009987},
    'GHG': {'ad': 'Dış Borçlanma Araçları Emeklilik Yatırım Fonu', 'agirlik': 0.20, 'giris_fiyati': 1.201487},
    'GHH': {'ad': 'Sürdürülebilirlik Hisse Senedi Emeklilik Yatırım Fonu', 'agirlik': 0.10, 'giris_fiyati': 0.395887}
}

tarih_str = datetime.now().strftime('%Y-%m-%d')
rapor_data = []

toplam_maliyet = 0.0
toplam_guncel_deger = 0.0
toplam_portfoye_katki = 0.0

tahmini_guncel_piyasa = {
    'GEL': 0.442970, 'GEH': 2.833800, 'EMY': 0.010250, 'GHG': 1.220412, 'GHH': 0.401200
}

for kod, info in fon_tanimlari.items():
    giris_fiy = float(info['giris_fiyati'])
    guncel_fiy = float(tahmini_guncel_piyasa.get(kod, giris_fiy))
    
    toplam_degisim_orani = float(((guncel_fiy - giris_fiy) / giris_fiy) * 100)
    fon_maliyeti = float(toplam_anapara * info['agirlik'])
    fon_guncel_degeri = float(fon_maliyeti * (1 + (toplam_degisim_orani / 100)))
    fon_net_kar_zarar = float(fon_guncel_degeri - fon_maliyeti)
    fon_portfoye_katki = float(info['agirlik'] * toplam_degisim_orani)
    
    toplam_maliyet += fon_maliyeti
    toplam_guncel_deger += fon_guncel_degeri
    toplam_portfoye_katki += fon_portfoye_katki
    
    rapor_data.append([
        kod, info['ad'], float(info['agirlik'] * 100), fon_maliyeti,
        giris_fiy, guncel_fiy, toplam_degisim_orani, fon_portfoye_katki, fon_net_kar_zarar
    ])

sutunlar = [
    'Fon Kodu', 'Fon Adı', 'Ağırlık (%)', 'Yatırılan Tutar (TL)', 
    'Giriş Fiyatı (TL)', 'Güncel Fiyat (TL)', 'Toplam Değişim (%)', 
    'Portföye Katkı (%)', 'Net Kâr/Zarar (TL)'
]
df = pd.DataFrame(rapor_data, columns=sutunlar)

genel_kar = float(toplam_guncel_deger - toplam_maliyet)
dogru_genel_degisim = float((genel_kar / toplam_maliyet) * 100)

toplam_satiri = pd.DataFrame([[
    'TOPLAM', 'Genel Portföy Durumu', 100.0, toplam_maliyet,
    None, None, dogru_genel_degisim, toplam_portfoye_katki, genel_kar
]], columns=sutunlar)

df = pd.concat([df, toplam_satiri], ignore_index=True)

excel_adi = f"BES_Raporu_{tarih_str}.xlsx"

with pd.ExcelWriter(excel_adi, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='BES Takip')
    workbook = writer.book
    worksheet = writer.sheets['BES Takip']
    
    # Formatlama
    for row in range(2, worksheet.max_row + 1):
        for col in range(3, 10):
            cell = worksheet.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = '0.000'

    # Sütun Genişlik Ayarı
    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col.column)
        for cell in col:
            if cell.value:
                val_str = f"{cell.value:.3f}" if isinstance(cell.value, float) else str(cell.value)
                max_len = max(max_len, len(val_str))
        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Grafik Entegrasyonu
    chart_bar = BarChart()
    chart_bar.type = "col"
    chart_bar.style = 11
    chart_bar.title = "Fon Bazlı Net Kâr / Zarar Detayı (TL)"
    chart_bar.width = 18   
    chart_bar.height = 10  
    data_bar = Reference(worksheet, min_col=9, min_row=1, max_row=6)
    cats_bar = Reference(worksheet, min_col=1, min_row=2, max_row=6)
    chart_bar.add_data(data_bar, titles_from_data=True)
    chart_bar.set_categories(cats_bar)
    chart_bar.legend = None 
    chart_bar.dataLabels = DataLabelList()
    chart_bar.dataLabels.showVal = True
    worksheet.add_chart(chart_bar, "A9")

print("Excel başarıyla oluşturuldu. E-posta gönderimi başlatılıyor...")

# INTERNAL PYTHON MAIL SENDER ENGINE
try:
    mail_user = os.environ.get('MAIL_USER')
    mail_pass = os.environ.get('MAIL_PASS')

    if mail_user and mail_pass:
        msg = MIMEMultipart()
        msg['From'] = f"BES Otomasyon Sistemi <{mail_user}>"
        msg['To'] = mail_user
        msg['Subject'] = f"Günlük Otomatik BES Durum Raporu - {tarih_str}"

        # Dosya Eki
        with open(excel_adi, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename= {excel_adi}")
            msg.attach(part)

        # SMTP Bağlantısı
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(mail_user, mail_pass)
        server.sendmail(mail_user, mail_user, msg.as_string())
        server.quit()
        print("E-posta başarıyla kutunuza gönderildi.")
    else:
        print("Hata: Gizli kasadaki (Secrets) mail kullanıcı bilgileri okunamadı.")
except Exception as e:
    print(f"E-posta Gönderim Hatası: {e}")
